import glob
import pandas as pd
from datetime import datetime, timedelta
import os
import numpy as np
from openpyxl import Workbook, load_workbook
from lottery_predictor import LotteryPredictor
from collections import Counter
from sklearn.cluster import KMeans
from data_analysis import DataAnalysis
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config.paths import PATHS, ensure_directories
from data_collector_selenium import KinoDataCollector
from prediction_evaluator import PredictionEvaluator
import joblib

class DrawHandler:
    def __init__(self):
        # Initialize paths using config
        ensure_directories()
        self.csv_file = PATHS['HISTORICAL_DATA']
        self.models_dir = PATHS['MODELS_DIR']
        self.predictions_dir = os.path.dirname(PATHS['PREDICTIONS'])
        
        # Initialize number columns
        self.number_cols = [f'number{i}' for i in range(1, 21)]
        
        # Pipeline status tracking remains the same
        self.pipeline_status = {
            'success': False,
            'stage': None,
            'error': None,
            'timestamp': None
        }
        
        # Initialize predictor
        self.predictor = LotteryPredictor()

        # Initialize continuous learning tracking
        self.learning_dir = os.path.join(self.models_dir, 'learning_history')
        os.makedirs(self.learning_dir, exist_ok=True)
        self.learning_history_file = os.path.join(self.learning_dir, 'learning_history.csv')
        self.learning_status = {
            'last_learning': None,
            'cycles_completed': 0,
            'initial_accuracy': None,
            'current_accuracy': None,
            'improvement_rate': None,
            'last_adjustments': []
        }
        self._load_learning_status()

    def handle_prediction_pipeline(self, historical_data=None):
        """Coordinates the prediction pipeline process"""
        try:
            self.pipeline_status['timestamp'] = datetime.now()
            
            if historical_data is None:
                historical_data = self._load_historical_data()
            
            # 1. Data Preparation Stage
            self.pipeline_status['stage'] = 'data_preparation'
            processed_data = self._prepare_pipeline_data(historical_data)
            
            # 2. Analysis Stage
            self.pipeline_status['stage'] = 'analysis'
            analysis_results = {}  # Initialize empty analysis results
            
            # 3. Prediction Stage
            self.pipeline_status['stage'] = 'prediction'
            model_path = self._get_latest_model()
            if model_path:
                model_files = [
                    f"{model_path}_prob_model.pkl",
                    f"{model_path}_pattern_model.pkl",
                    f"{model_path}_scaler.pkl"
                ]
                if all(os.path.exists(file) for file in model_files):
                    print(f"✓ Model found: {os.path.basename(model_path)}")
                    predictions, probabilities, analysis_results = self._run_prediction(processed_data)
                    if predictions is not None:
                        self.pipeline_status['success'] = True
                        return predictions, probabilities, analysis_results
                else:
                    print("Model files incomplete. Attempting retraining...")
                    if self.train_ml_models():
                        # After training, run prediction again with the new model
                        return self._run_prediction(processed_data)
            else:
                print("No model found. Attempting training...")
                if self.train_ml_models():
                    # After training, run prediction with the new model
                    return self._run_prediction(processed_data)
            
            # If we reach here, prediction failed
            self.pipeline_status['error'] = "Failed to generate predictions"
            return None, None, None
                
        except Exception as e:
            print(f"Error in prediction pipeline: {e}")
            self.pipeline_status['error'] = str(e)
            self.pipeline_status['success'] = False
            return None, None, None

    def train_ml_models(self, force_retrain=False):
        """Train or retrain ML models"""
        try:
            self.pipeline_status['stage'] = 'model_training'
            
            # Load historical data
            historical_data = self._load_historical_data()
            if historical_data is None or len(historical_data) < 6:
                raise ValueError("Insufficient historical data for training")

            print(f"Loaded {len(historical_data)} draws for training")
                
            # Prepare data for training
            features, labels = self.predictor.prepare_data(historical_data)
            if features is None or labels is None or len(features) == 0:
                raise ValueError("Failed to prepare training data")
                
            print(f"Prepared features shape: {features.shape}")

            # Train models using predictor with prepared data
            training_success = self.predictor.train_models(features, labels)
            
            if training_success:
                # Save models immediately after successful training
                if not self.predictor.save_models():
                    raise Exception("Failed to save trained models")
                    
                self.pipeline_status['success'] = True
                print("Models trained and saved successfully")
                return True
            else:
                raise Exception("Model training failed")
                
        except Exception as e:
            self.pipeline_status['error'] = str(e)
            print(f"Error in model training: {e}")
            return False

    def save_draw_to_csv(self, draw_date, draw_numbers, csv_file=None):
        if csv_file is None:
            csv_file = self.csv_file
        return save_draw_to_csv(draw_date, draw_numbers, csv_file)

    def save_predictions_to_csv(self, predicted_numbers, probabilities, timestamp, csv_file=None):
        if csv_file is None:
            csv_file = PATHS['PREDICTIONS']
        return save_predictions_to_csv(predicted_numbers, probabilities, timestamp, csv_file)

    def save_predictions_to_excel(self, predictions, probabilities, timestamp, excel_file=None):
        if excel_file is None:
            excel_file = PATHS['ANALYSIS']
        return save_predictions_to_excel(predictions, probabilities, timestamp, excel_file)
    
    def _get_latest_model(self):
        """Get the path to the latest model"""
        try:
            # First check in the models directory for timestamp file
            timestamp_file = os.path.join(self.models_dir, 'model_timestamp.txt')
            if os.path.exists(timestamp_file):
                with open(timestamp_file, 'r') as f:
                    timestamp = f.read().strip()
                    model_path = os.path.join(self.models_dir, f'lottery_predictor_{timestamp}')
                    if os.path.exists(f"{model_path}_prob_model.pkl"):
                        return model_path
            
            # Fallback to searching for model files
            model_files = glob.glob(os.path.join(self.models_dir, "*_prob_model.pkl"))
            if model_files:
                latest = max(model_files, key=os.path.getctime)
                return latest.replace('_prob_model.pkl', '')
                
            return None
                
        except Exception as e:
            print(f"Error getting latest model: {e}")
            return None

    def _load_historical_data(self):
        """Load historical data from CSV"""
        try:
            return load_data(self.csv_file)
        except Exception as e:
            print(f"Error loading historical data: {e}")
            return None

    def _prepare_pipeline_data(self, data):
        """Prepare data for prediction pipeline"""
        try:
            if data is None:
                return None
            
            # Basic data cleaning
            data = data.copy()
            data = data.dropna(subset=['date'])
            
            # Extract date features
            data = extract_date_features(data)
            
            # Ensure all number columns exist
            for col in self.number_cols:
                if col not in data.columns:
                    data[col] = 0
                    
            return data
            
        except Exception as e:
            print(f"Error preparing data: {e}")
            return None

    def _run_prediction(self, data):
        """Run prediction with analysis integration"""
        try:
            analysis_results = {}  # Initialize empty analysis results
            
            # Use the existing predictor instance instead of creating a new one
            model_base = self._get_latest_model()
            
            if not model_base:
                raise ValueError("No model base found. Training required.")
                
            # Try loading models
            if not self.predictor.load_models(model_base):
                print("Model loading failed, attempting to train new model...")
                if not self.train_ml_models():
                    raise ValueError("Could not load or train models")
                # Get new model path after training
                model_base = self._get_latest_model()
                if not model_base or not self.predictor.load_models(model_base):
                    raise ValueError("Model loading failed after training")
            
            number_cols = [f'number{i}' for i in range(1, 21)]
            recent_draws = data.tail(5)[number_cols + ['date', 'day_of_week', 'month', 'day_of_year', 'days_since_first_draw']]
            
            # Get predictions
            predicted_numbers, probabilities, analysis = self.predictor.predict(recent_draws)
            
            # Update analysis results
            if analysis:
                analysis_results.update(analysis)
                if self.predictor.training_status.get('prob_score') is not None:
                    analysis_results['model_performance'] = {
                        'probabilistic_model_score': self.predictor.training_status['prob_score'],
                        'pattern_model_score': self.predictor.training_status['pattern_score']
                    }
            
            return predicted_numbers, probabilities, analysis_results
                
        except Exception as e:
            print(f"Error in prediction run: {e}")
            return None, None, None

    def _handle_pipeline_results(self, predictions, probabilities, analysis_results):
        """Handle the results from the prediction pipeline"""
        try:
            if predictions is None or probabilities is None:
                print("No valid predictions to handle")
                return False
                
            # Format timestamp and next draw time
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            next_draw_time = get_next_draw_time(datetime.now())
            
            # Save predictions to CSV
            self.save_predictions_to_csv(predictions, probabilities, timestamp)
            
            # Also save to Excel for better visualization
            self.save_predictions_to_excel(predictions, probabilities, timestamp)
            
            # Display formatted results
            formatted_numbers = ','.join(map(str, sorted(predictions)))
            print(f"\nPredicted numbers for next draw at {next_draw_time.strftime('%H:%M %d-%m-%Y')}:")
            print(f"Numbers: {formatted_numbers}")
            
            # Display probabilities
            print("\nProbabilities for each predicted number:")
            for num, prob in zip(sorted(predictions), 
                               [probabilities[num - 1] for num in predictions]):
                print(f"Number {num}: {prob:.4f}")
            
            # Save analysis results if available
            if analysis_results:
                # Save hot numbers if available
                if 'hot_numbers' in analysis_results:
                    top_4_numbers = analysis_results['hot_numbers'][:4]
                    top_4_file_path = os.path.join(self.predictions_dir, 'top_4.xlsx')
                    save_top_4_numbers_to_excel(top_4_numbers, top_4_file_path)
                    print(f"\nTop 4 numbers based on analysis: {','.join(map(str, top_4_numbers))}")
                
                # Display analysis summary
                print("\n=== Analysis Results ===")
                for key, value in analysis_results.items():
                    if key != 'clusters':  # Skip clusters for cleaner output
                        print(f"\n{key.replace('_', ' ').title()}:")
                        print(value)
                        
            return True
        except Exception as e:
            print(f"Error handling pipeline results: {e}")
            return False

    # NEW METHODS FOR CONTINUOUS LEARNING
    
    def _load_learning_status(self):
        """Load current learning status from history file"""
        try:
            if os.path.exists(self.learning_history_file):
                df = pd.read_csv(self.learning_history_file)
                if not df.empty:
                    last_row = df.iloc[-1]
                    
                    self.learning_status['last_learning'] = last_row['timestamp']
                    self.learning_status['cycles_completed'] = len(df)
                    
                    if len(df) > 1:
                        self.learning_status['initial_accuracy'] = df.iloc[0]['accuracy']
                        self.learning_status['current_accuracy'] = last_row['accuracy']
                        self.learning_status['improvement_rate'] = (
                            (last_row['accuracy'] - df.iloc[0]['accuracy']) / df.iloc[0]['accuracy'] * 100
                            if df.iloc[0]['accuracy'] > 0 else 0
                        )
                        
                        # Get last adjustments if available
                        if 'adjustments' in last_row:
                            try:
                                self.learning_status['last_adjustments'] = eval(last_row['adjustments'])
                            except:
                                pass
                        
                    print(f"Loaded learning history: {self.learning_status['cycles_completed']} learning cycles completed")
                    
        except Exception as e:
            print(f"Error loading learning status: {e}")
    
    def apply_learning_from_evaluations(self):
        """
        Apply continuous learning by analyzing evaluation results
        and making adjustments to the prediction model.
        """
        try:
            print("\nApplying continuous learning from evaluation results...")
            
            # Initialize the evaluator
            evaluator = PredictionEvaluator()
            
            # Get performance statistics
            stats = evaluator.get_performance_stats()
            if not stats or stats.get('total_predictions', 0) < 5:
                print("Not enough evaluation data for learning (need at least 5 predictions)")
                return False
                
            # Extract insights from evaluation results
            problematic_numbers = list(stats.get('most_frequently_missed', {}).keys())
            successful_numbers = list(stats.get('most_frequent_correct', {}).keys())
            recent_trend = stats.get('recent_trend', 0)
            average_accuracy = stats.get('average_accuracy', 0)
            
            print(f"\nEvaluation insights:")
            print(f"- Problematic numbers: {problematic_numbers}")
            print(f"- Successful numbers: {successful_numbers}")
            print(f"- Recent trend: {recent_trend:.3f} ({'improving' if recent_trend > 0 else 'declining'})")
            print(f"- Average accuracy: {average_accuracy:.2f}%")
            
            # Create adjustment plan
            adjustments = {
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'adjustments_made': []
            }
            
            # Make appropriate adjustments to model based on insights
            self._adjust_model_parameters(problematic_numbers, successful_numbers, recent_trend, average_accuracy, adjustments)
            
            # Save learning metadata
            self._save_learning_metadata(stats, adjustments)
            
            # Update learning status
                       # Update learning status
            self.learning_status['last_learning'] = adjustments['timestamp']
            self.learning_status['cycles_completed'] += 1
            self.learning_status['current_accuracy'] = average_accuracy
            
            if self.learning_status['initial_accuracy'] is None:
                self.learning_status['initial_accuracy'] = average_accuracy
                
            self.learning_status['improvement_rate'] = (
                (average_accuracy - self.learning_status['initial_accuracy']) / self.learning_status['initial_accuracy'] * 100
                if self.learning_status['initial_accuracy'] and self.learning_status['initial_accuracy'] > 0 else 0
            )
            
            self.learning_status['last_adjustments'] = adjustments['adjustments_made']
            
            print("\nLearning cycle completed:")
            print(f"- Total learning cycles: {self.learning_status['cycles_completed']}")
            print(f"- Current accuracy: {self.learning_status['current_accuracy']:.2f}%")
            print(f"- Total improvement: {self.learning_status['improvement_rate']:.2f}%")
            
            return True
            
        except Exception as e:
            print(f"Error applying learning: {e}")
            return False
    
    def _adjust_model_parameters(self, problematic_numbers, successful_numbers, trend, accuracy, adjustments):
        """Make specific adjustments to model parameters based on insights"""
        try:
            # Define timestamp here - at the TOP of the method
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Ensure predictor is initialized
            if not hasattr(self.predictor, 'pipeline_data'):
                self.predictor.pipeline_data = {}
            
            # 1. Create number boosts for problematic numbers
            if problematic_numbers:
                boost_array = np.ones(80)
                for num in problematic_numbers[:10]:  # Take top 10 most missed
                    if 1 <= num <= 80:
                        boost_array[num-1] = 1.05  # 5% boost for missed numbers
            
                self.predictor.pipeline_data['number_boosts'] = boost_array
                adjustments['adjustments_made'].append(f"Boosted problematic numbers: {problematic_numbers[:10]}")
            
            # 2. Adjust model weights based on trend
            if trend < 0:  # If performance is declining
                # Modify weights between probability and pattern models
                weights = {
                    'prob_weight': 0.5,  # Default is 0.4
                    'pattern_weight': 0.5  # Default is 0.6
                }
                self.predictor.pipeline_data['prediction_weights'] = weights
                adjustments['adjustments_made'].append("Modified model weights due to declining trend")
            
            # 3. If accuracy is very low, make more significant changes
            if accuracy < 10:  # Below 10% accuracy
                # Enable enhanced feature extraction
                self.predictor.pipeline_data['use_enhanced_features'] = True
                
                # Consider adjusting the training data balance
                if hasattr(self.predictor, 'training_status'):
                    self.predictor.training_status['require_retraining'] = True
                
                adjustments['adjustments_made'].append("Enabled enhanced features due to low accuracy")
            
            # 4. Save a new model if adjustments were made
            if adjustments['adjustments_made']:
                model_path = os.path.join(self.models_dir, f'lottery_predictor_adjusted_{timestamp}')
                
                # Save the adjusted model
                self.predictor.save_models(path_prefix=model_path)
                
                # Save adjustment metadata
                adjustment_file = os.path.join(self.models_dir, 'model_adjustments.txt')
                with open(adjustment_file, 'a') as f:
                    f.write(f"\n--- Adjustments made at {adjustments['timestamp']} ---\n")
                    for adjustment in adjustments['adjustments_made']:
                        f.write(f"- {adjustment}\n")
                
                print(f"Model adjusted and saved as: {os.path.basename(model_path)}")
                
                # Update timestamp file to point to the new adjusted model
                timestamp_file = os.path.join(self.models_dir, 'model_timestamp.txt')
                with open(timestamp_file, 'w') as f:
                    f.write(timestamp)
                
                return True
            else:
                print("No adjustments needed at this time")
                return False
                
        except Exception as e:
            print(f"Error adjusting model parameters: {e}")
            return False
    
    def _save_learning_metadata(self, stats, adjustments):
        """Save metadata about the learning process"""
        try:
            # Create or append to learning history CSV
            data = {
                'timestamp': [adjustments['timestamp']],
                'accuracy': [stats.get('average_accuracy', 0)],
                'trend': [stats.get('recent_trend', 0)],
                'best_prediction': [stats.get('best_prediction', 0)],
                'adjustments': [str(adjustments['adjustments_made'])]
            }
            
            df = pd.DataFrame(data)
            
            if os.path.exists(self.learning_history_file):
                mode = 'a'
                header = False
            else:
                mode = 'w'
                header = True
                
            df.to_csv(self.learning_history_file, mode=mode, header=header, index=False)
            print(f"Learning metadata saved to {self.learning_history_file}")
            
        except Exception as e:
            print(f"Error saving learning metadata: {e}")
    
    def run_continuous_learning_cycle(self):
        """Run a complete continuous learning cycle"""
        try:
            print("\nStarting continuous learning cycle...")
            
            # Initialize pipeline tracking at the beginning of the method
            self.pipeline_tracking = {
                'start_time': datetime.now(),
                'stages_completed': [],
                'current_stage': None,
                'error': None
            }
            
            # Step 1: Evaluate past predictions
            print("\nStep 1: Evaluating past predictions...")
            evaluator = PredictionEvaluator()
            evaluator.evaluate_past_predictions()
            
            # Step 2: Apply learning from evaluations
            print("\nStep 2: Applying learning from evaluations...")
            learning_success = self.apply_learning_from_evaluations()
            
            # Step 3: Test the improved model
            if learning_success:
                print("\nStep 3: Testing improved model...")
                
                # Reset predictor status to ensure fresh initialization
                if hasattr(self.predictor, 'training_status'):
                    self.predictor.training_status['model_loaded'] = False
                    
                try:
                    # Force model reload
                    model_path = self._get_latest_model()
                    if model_path:
                        print(f"Loading adjusted model: {os.path.basename(model_path)}")
                        load_success = self.predictor.load_models(model_path)
                        
                        if load_success:
                            print("Model loaded successfully, generating new prediction...")
                            
                            # Initialize a fresh pipeline tracking dictionary for prediction test
                            test_pipeline_tracking = {
                                'start_time': datetime.now(),
                                'stages_completed': [],
                                'current_stage': None,
                                'error': None
                            }
                            
                            # Store original pipeline_tracking
                            original_tracking = self.pipeline_tracking
                            
                            # Set the test tracking for prediction
                            self.pipeline_tracking = test_pipeline_tracking
                            
                            try:
                                # Generate prediction with reloaded model
                                predictions, probabilities, analysis = self.handle_prediction_pipeline()
                                if predictions is not None:
                                    print(f"\nGenerated prediction with improved model: {sorted(predictions)}")
                                else:
                                    print("Failed to generate prediction with improved model")
                            finally:
                                # Restore original pipeline tracking
                                self.pipeline_tracking = original_tracking
                        else:
                            print("Failed to load adjusted model")
                    else:
                        print("No adjusted model found")
                except Exception as e:
                    print(f"Error testing improved model: {e}")
                    
            print("\nContinuous learning cycle complete!")
            return True
                
        except Exception as e:
            print(f"Error in continuous learning cycle: {e}")
            return False
            
    def get_learning_metrics(self):
        """Get current learning metrics for display"""
        return {
            'cycles_completed': self.learning_status['cycles_completed'],
            'last_learning': self.learning_status['last_learning'],
            'initial_accuracy': self.learning_status['initial_accuracy'],
            'current_accuracy': self.learning_status['current_accuracy'],
            'improvement_rate': self.learning_status['improvement_rate'],
            'last_adjustments': self.learning_status['last_adjustments']
        }

def save_draw_to_csv(draw_date, draw_numbers, csv_file=None):
    """Save draw results to CSV"""
    if csv_file is None:
        csv_file = PATHS['HISTORICAL_DATA']
    try:
        # Prepare data
        draw_data = {
            'date': [draw_date],
            **{f'number{i+1}': [num] for i, num in enumerate(sorted(draw_numbers))}
        }
        df = pd.DataFrame(draw_data)
        
        # Save to CSV
        os.makedirs(os.path.dirname(csv_file), exist_ok=True)
        if os.path.exists(csv_file):
            df.to_csv(csv_file, mode='a', header=False, index=False)
        else:
            df.to_csv(csv_file, index=False)
        return True
    except Exception as e:
        print(f"Error saving draw to CSV: {e}")
        return False

def save_predictions_to_csv(predicted_numbers, probabilities, timestamp, csv_file=None):
    """Save predictions to CSV"""
    if csv_file is None:
        csv_file = PATHS['PREDICTIONS']
    try:
        data = {
            'Timestamp': [timestamp],
            'Predicted_Numbers': [','.join(map(str, predicted_numbers))],
            'Probabilities': [','.join(map(str, [probabilities[num - 1] for num in predicted_numbers]))]
        }
        df = pd.DataFrame(data)
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(csv_file), exist_ok=True)
        
        if os.path.exists(csv_file):
            df.to_csv(csv_file, mode='a', header=False, index=False)
        else:
            df.to_csv(csv_file, index=False)
        return True
    except Exception as e:
        print(f"Error saving predictions to CSV: {e}")
        return False

def save_predictions_to_excel(predictions, probabilities, timestamp, excel_file=None):
    """Save predictions to Excel"""
    if excel_file is None:
        excel_file = PATHS['ANALYSIS']
    try:
        data = {
            'Timestamp': [timestamp],
            'Predicted_Numbers': [','.join(map(str, predictions))],
            'Probabilities': [','.join(map(str, [probabilities[num - 1] for num in predictions]))]
        }
        df = pd.DataFrame(data)
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(excel_file), exist_ok=True)
        
        if os.path.exists(excel_file):
            # Load existing workbook
            book = load_workbook(excel_file)
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row + 1)
        else:
            # Create new workbook
            df.to_excel(excel_file, index=False)
        return True
    except Exception as e:
        print(f"Error saving predictions to Excel: {e}")
        return False

def load_data(file_path=None):
    """Load and preprocess data from CSV"""
    if file_path is None:
        file_path = PATHS['HISTORICAL_DATA']
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Data file {file_path} not found")
        
    df = pd.read_csv(file_path)
    number_cols = [f'number{i}' for i in range(1, 21)]  # Define number_cols here
    
    try:
        df['date'] = pd.to_datetime(df['date'], format='%H:%M %d-%m-%Y', errors='coerce')
        df.loc[df['date'].isna(), 'date'] = pd.to_datetime(df.loc[df['date'].isna(), 'date'], errors='coerce')
    except Exception as e:
        print(f"Warning: Date conversion issue: {e}")
        
    try:
        df[number_cols] = df[number_cols].astype(float)
    except Exception as e:
        print(f"Warning: Could not process number columns: {e}")
        
    for col in number_cols:
        if col in df.columns:
            df[col] = df[col].fillna(df[col].mode()[0] if not df[col].mode().empty else 0)
            
    return df

def extract_date_features(df):
    df['day_of_week'] = df['date'].dt.dayofweek
    df['hour'] = df['date'].dt.hour
    df['minute'] = df['date'].dt.minute
    df['month'] = df['date'].dt.month
    df['day_of_year'] = df['date'].dt.dayofyear
    df['days_since_first_draw'] = (df['date'] - df['date'].min()).dt.days
    return df

def get_next_draw_time(current_time):
    minutes = (current_time.minute // 5 + 1) * 5
    next_draw_time = current_time.replace(minute=0, second=0, microsecond=0) + timedelta(minutes=minutes)
    return next_draw_time

def save_top_4_numbers_to_excel(top_4_numbers, file_path=None):
    if file_path is None:
        file_path = os.path.join(os.path.dirname(PATHS['PREDICTIONS']), 'top_4.xlsx')
    df = pd.DataFrame({'Top 4 Numbers': top_4_numbers})
    
    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    df.to_excel(file_path, index=False)

def evaluate_numbers(historical_data):
    """Evaluate numbers based on criteria other than frequency.
    For simplicity, this example assumes a dummy evaluation function.
    Replace this with your actual evaluation logic.
    """
    number_evaluation = {i: 0 for i in range(1, 81)}
    for index, row in historical_data.iterrows():
        for i in range(1, 21):
            number = row[f'number{i}']
            number_evaluation[number] += 1  # Dummy evaluation logic
    sorted_numbers = sorted(number_evaluation, key=number_evaluation.get, reverse=True)
    return sorted_numbers[:4]  # Return top 4 numbers

def train_and_predict():
    try:
        handler = DrawHandler()    
        
        # First ensure models are trained
        print("Checking/Training models...")
        if not handler.train_ml_models():
            raise Exception("Model training failed")
        print("Models ready")
        
        # Generate prediction using pipeline
        print("\nGenerating predictions...")
        predictions, probabilities, analysis = handler.handle_prediction_pipeline()
        
        if predictions is not None:
            # Format and display predictions
            formatted_numbers = ','.join(map(str, sorted(predictions)))
            next_draw_time = get_next_draw_time(datetime.now())
            print(f"\nPredicted numbers for next draw at {next_draw_time.strftime('%H:%M %d-%m-%Y')}:")
            print(f"Numbers: {formatted_numbers}")
            
            # Display probabilities in a readable format
            print("\nProbabilities for each predicted number:")
            for num, prob in zip(sorted(predictions), 
                               [probabilities[num - 1] for num in predictions]):
                print(f"Number {num}: {prob:.4f}")

            # Save predictions
            handler.save_predictions_to_csv(predictions, probabilities, 
                                         next_draw_time.strftime('%Y-%m-%d %H:%M:%S'))
            
            # Handle top 4 numbers if available
            if analysis and 'hot_numbers' in analysis:
                top_4_numbers = analysis['hot_numbers'][:4]
                top_4_file_path = os.path.join(os.path.dirname(PATHS['PREDICTIONS']), 'top_4.xlsx')
                save_top_4_numbers_to_excel(top_4_numbers, top_4_file_path)
                print(f"\nTop 4 numbers based on analysis: {','.join(map(str, top_4_numbers))}")
            
            return predictions, probabilities, analysis
        else:
            print("\nFailed to generate predictions")
            return None, None, None
            
    except Exception as e:
        print(f"\nError in prediction process: {str(e)}")
        return None, None, None

def perform_complete_analysis(draws):
    """Perform all analyses and save to Excel"""
    try:
        if not draws:
            collector = KinoDataCollector()    
            draws = collector.fetch_latest_draws()
        
        if draws:
            analysis = DataAnalysis(draws)    
            
            # Perform all analyses
            analysis_data = {
                'frequency': analysis.get_top_numbers(20),
                'suggested_numbers': analysis.suggest_numbers(),
                'common_pairs': analysis.find_common_pairs(),
                'consecutive_numbers': analysis.find_consecutive_numbers(),
                'range_analysis': analysis.number_range_analysis(),
                'hot_cold_numbers': analysis.hot_and_cold_numbers()
            }
            
            # Save to Excel file using config path
            excel_path = PATHS['ANALYSIS']
            os.makedirs(os.path.dirname(excel_path), exist_ok=True)
            
            # Create Excel writer
            with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
                # Frequency Analysis
                pd.DataFrame({
                    'Top 20 Numbers': analysis_data['frequency'],
                    'Frequency Count': [analysis.count_frequency().get(num, 0) for num in analysis_data['frequency']]
                }).to_excel(writer, sheet_name='Frequency Analysis', index=False)
                
                # Suggested Numbers
                pd.DataFrame({
                    'Suggested Numbers': analysis_data['suggested_numbers']
                }).to_excel(writer, sheet_name='Suggested Numbers', index=False)
                
                # Common Pairs
                pd.DataFrame(analysis_data['common_pairs'], 
                           columns=['Pair', 'Frequency']
                ).to_excel(writer, sheet_name='Common Pairs', index=False)
                
                # Consecutive Numbers
                pd.DataFrame({
                    'Consecutive Sets': [str(x) for x in analysis_data['consecutive_numbers']]
                }).to_excel(writer, sheet_name='Consecutive Numbers', index=False)
                
                # Range Analysis
                pd.DataFrame(analysis_data['range_analysis'].items(),
                           columns=['Range', 'Count']
                ).to_excel(writer, sheet_name='Range Analysis', index=False)
                
                # Hot and Cold Numbers
                hot, cold = analysis_data['hot_cold_numbers']
                pd.DataFrame({
                    'Hot Numbers': hot,
                    'Cold Numbers': cold
                }).to_excel(writer, sheet_name='Hot Cold Analysis', index=False)
            
            print(f"\nComplete analysis saved to: {excel_path}")
            return True
    except Exception as e:
        print(f"\nError in complete analysis: {str(e)}")
        return False

def test_pipeline_integration():
    """Test the integrated prediction pipeline"""
    try:
        handler = DrawHandler()
        pipeline_status = {
            'data_collection': False,
            'analysis': False,
            'prediction': False,
            'evaluation': False
        }
        
        # 1. Data Collection
        print("\nStep 1: Collecting data...")
        collector = KinoDataCollector()
        draws = collector.fetch_latest_draws()
        if draws:
            pipeline_status['data_collection'] = True
            print("✓ Data collection successful")
            
            # Save draws to CSV
            for draw_date, numbers in draws:
                handler.save_draw_to_csv(draw_date, numbers)

            # 2. Analysis
            print("\nStep 2: Performing analysis...")
            if perform_complete_analysis(draws):
                pipeline_status['analysis'] = True
                print("✓ Analysis complete and saved")

            # 3. ML Prediction
            print("\nStep 3: Generating prediction...")
            predictions, probabilities, analysis = handler.handle_prediction_pipeline()
            if predictions is not None:
                pipeline_status['prediction'] = True
                print("✓ Prediction generated")
                # Display prediction results
                formatted_numbers = ','.join(map(str, predictions))
                print(f"Predicted numbers: {formatted_numbers}")
                if analysis and 'hot_numbers' in analysis:
                    print(f"Hot numbers: {analysis['hot_numbers'][:10]}")

            # 4. Evaluation
            print("\nStep 4: Evaluating predictions...")
            evaluator = PredictionEvaluator()
            evaluator.evaluate_past_predictions()
            pipeline_status['evaluation'] = True
            print("✓ Evaluation complete")
        
        return pipeline_status

    except Exception as e:
        print(f"\nError in pipeline: {str(e)}")
        return pipeline_status

def main():
    handler = DrawHandler()
    
    try:
        # First ensure models are trained
        print("Checking/Training models...")
        if handler.train_ml_models():
            print("Models ready")
        
        # Generate prediction using pipeline
        print("\nGenerating predictions...")
        predictions, probabilities, analysis = handler.handle_prediction_pipeline()
        
        if predictions is not None:
            print("\n=== Prediction Results ===")
            print(f"Predicted Numbers: {sorted(predictions)}")
            
            # Display probabilities for predicted numbers
            print("\nProbabilities for predicted numbers:")
            for num, prob in zip(sorted(predictions), 
                               [probabilities[num-1] for num in predictions]):
                print(f"Number {num}: {prob:.4f}")
            
            if analysis:
                print("\n=== Analysis Results ===")
                for key, value in analysis.items():
                    if key != 'clusters':  # Skip clusters for cleaner output
                        print(f"\n{key.replace('_', ' ').title()}:")
                        print(value)
                        
    except Exception as e:
        print(f"\nError in pipeline execution: {str(e)}")
        print(f"Pipeline Status: {handler.pipeline_status}")